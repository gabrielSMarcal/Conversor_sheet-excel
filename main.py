from datetime import datetime

import openpyxl
import unidecode
import csv


# Load the workbook
def load_workbook(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        return wb
    except Exception as e:
        print(f"An error occurred while loading the workbook: {e}")
        return None


# Get the header
def get_header(sheet):
    max_column = sheet.max_column
    row_value = 1
    header = []
    for col in range(1, max_column + 1):
        cell_value = sheet.cell(row=row_value, column=col).value
        if cell_value is None:
            break
        header_value = unidecode.unidecode(cell_value.lower().replace(" ", "_").replace("/", "_"))
        header.append(header_value)
    return header


# Read the content
def read_content(sheet, header):
    min_row = 2
    table = []
    for row in sheet.iter_rows(min_row=min_row, max_row=None, min_col=1, max_col=len(header)):
        row_data = {}
        for col_idx, cell in enumerate(row):
            if col_idx < len(header):
                header_name = header[col_idx]
                if header_name == 'data' and cell.value is not None:
                    if isinstance(cell.value, datetime):
                        row_data[header[col_idx]] = cell.value.strftime('%d/%m/%Y')
                    else:
                        try:
                            date_value = datetime.strptime(cell.value, '%Y-%m-%d')
                            row_data[header[col_idx]] = date_value.strftime('%d/%m/%Y')
                        except ValueError:
                            row_data[header[col_idx]] = cell.value
                elif header_name == 'cpf' and cell.value is not None:
                    row_data[header[col_idx]] = cell.value.replace(".", "").replace("-", "")
                else:
                    row_data[header[col_idx]] = cell.value

        if any(value is not None for value in row_data.values()):
            table.append(row_data)
    return table

# Write to CSV
def write_to_csv(table, header, output_csv_path):
    try:
        with open(output_csv_path, mode="w", newline="", encoding="utf-8") as csv_file:
            writer = csv.writer(csv_file, delimiter=";")
            writer.writerow(header)  # Header
            for row in table:
                writer.writerow([row.get(col, "") for col in header])  # Rows
    except Exception as e:
        print(f"An error occurred while writing to CSV: {e}")

# Read the Excel file
def read_excel_file(file_path, output_csv_path):
    wb = load_workbook(file_path)
    if wb is None:
        return None

    sheet = wb.active
    header = get_header(sheet)
    table = read_content(sheet, header)
    write_to_csv(table, header, output_csv_path)
    return table


def main():
    file_path = "teste.xlsx"
    output_csv_path = "output_teste.csv"
    read_excel_file(file_path, output_csv_path)

if __name__ == "__main__":
    main()